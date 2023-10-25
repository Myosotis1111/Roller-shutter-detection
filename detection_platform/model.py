import os
import os.path as osp
import shutil
import string
from datetime import datetime, timedelta
from pathlib import Path
import cv2
import numpy as np
import openpyxl
import torch
from PIL import Image
from PySide2.QtCore import QDateTime
from PySide2.QtGui import QImage, QPixmap
from PySide2.QtWidgets import QFileDialog, QMessageBox
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Border, Font, Side, Alignment
from openpyxl.utils import get_column_letter
from signal import MySignals
from utils.datasets import LoadImages
from utils.general import check_img_size, scale_coords, LOGGER, non_max_suppression
from utils.plots import Annotator, colors
from utils.torch_utils import select_device, time_sync
from window import Window
from detection_model_load import model_load


class Model:

    def __init__(self):

        self.ms = MySignals()
        self.mode = None  # mode = "GUI" or "test", the former updates GUI and the latter for evaluation does not.
        self.iou_thres = None
        self.conf_thres = None
        self.device = None
        self.output_size = None
        self.model = None
        self.window_info = None
        self.modified_status = None
        self.workbook = None
        self.output_size = 640
        self.img2predict = ""
        self.device = ''  # GPU or CPU will be automatically selected for detection.
        self.model = model_load(weights="weights/best.pt", device=self.device)  # load detection model and weights
        self.windows = {}
        self.window_no = None

    def initializeWindows(self):

        # initialize window instance
        for i in range(1, 28):
            window_no = f"upper_{i}"
            self.windows[window_no] = Window(doc=0, doc_min=0, doc_max=100, status="not_detected", count_open=0,
                                             count_closed=0)

        for i in range(1, 28):
            window_no = f"middle_{i}"
            self.windows[window_no] = Window(doc=0, doc_min=0, doc_max=100, status="not_detected", count_open=0,
                                             count_closed=0)

        for i in range(1, 19):
            window_no = f"lower_{i}"
            self.windows[window_no] = Window(doc=0, doc_min=0, doc_max=100, status="not_detected", count_open=0,
                                             count_closed=0)

    def upload(self):

        self.initializeWindows()

        # read images
        fileNames, _ = QFileDialog.getOpenFileNames(None, 'Choose files', '', '*.jpg *.png *.tif *.jpeg')
        if fileNames:

            # If more than 3 images are uploaded, display the warning
            if len(fileNames) > 3:
                QMessageBox.warning(None, 'Warning', 'Maximum three images can be uploaded at once!')
                return

            self.uploadPreprocessing(fileNames, self.mode)

    def uploadPreprocessing(self, filenames, mode):

        self.img2predict = []
        self.window_no = []
        last_upload_time = None

        # iterate every image uploaded
        for i, fileName in enumerate(filenames):
            suffix = fileName.split(".")[-1]
            if mode == "GUI":
                save_path = osp.join("images/tmp", f"tmp_upload_{i}." + suffix)
            if mode == "test":
                save_path = osp.join("evaluation/tmp", f"tmp_upload_{i}." + suffix)
            shutil.copy(fileName, save_path)
            im = cv2.imread(save_path)
            # Read the labeled txt file, note that it should have the same name as
            # the corresponding image and be placed in the same folder.
            txt_filename, ext = os.path.splitext(fileName)
            txt_path = f"{txt_filename}.txt"

            # If no text file found, display the warning.
            if not os.path.exists(txt_path):
                QMessageBox.warning(None, 'Warning', 'No txt file found!')
                return

            # Obtain photo taken time, usually the file name may contain the time,
            # like image_dingcam03_2023-05-16_04-00-06. If not, the recorded taken time in the file properties
            # will be used. This will also check if the images are taken within one hour to ensure the detection
            # is valid.
            if mode == "GUI":
                basename = os.path.basename(fileName)
                try:
                    nameWithoutExt = os.path.splitext(basename)[0]
                    printable = set(string.printable)
                    nameWithoutExt = ''.join(filter(lambda x: x in printable, nameWithoutExt))
                    second_underscore_index = nameWithoutExt.find('_', nameWithoutExt.find('_') + 1)
                    fourth_underscore_index = nameWithoutExt.find('_', nameWithoutExt.find('_',
                                                                                           second_underscore_index + 1) + 1)
                    date_time_str = nameWithoutExt[second_underscore_index + 1:fourth_underscore_index]
                    date_time = datetime.strptime(date_time_str, '%Y-%m-%d_%H-%M-%S')

                    if last_upload_time is None:
                        last_upload_time = date_time
                    if last_upload_time is not None and abs(date_time - last_upload_time) > timedelta(
                            hours=1):
                        QMessageBox.warning(None, 'Warning',
                                            'Make sure uploading images during one same period!')
                        return
                    last_upload_time = date_time

                    if mode == "GUI":
                        self.ms.update_dateTime.emit(date_time)

                # If the file name is not standard (contains photo taken time)
                except ValueError:

                    exif_data = Image.open(save_path)._getexif()
                    # initialize the time displayed
                    date_time = QDateTime(2000, 1, 1, 0, 0, 0)

                    if exif_data:
                        date_time = exif_data.get(36867)
                        if date_time:
                            date_time = datetime.strptime(date_time, '%Y:%m:%d %H:%M:%S')
                            if last_upload_time is None:
                                last_upload_time = date_time
                            if last_upload_time is not None and abs(date_time - last_upload_time) > timedelta(
                                    hours=1):
                                QMessageBox.warning(None, 'Warning',
                                                    'Make sure uploading images during one same period!')
                                return
                            last_upload_time = date_time

                    self.ms.update_dateTime.emit(date_time)

            # cropping images according to coordinates in the text file
            with open(txt_path, 'r') as file:
                index = 0
                for line in file:
                    parts = line.strip().split(' ')
                    self.window_no.append(str(parts[0]))
                    window_no = parts[0]
                    coord_upper_left = eval(parts[1])
                    coord2_upper_right = eval(parts[2])
                    coord3_lower_right = eval(parts[3])
                    coord4_lower_left = eval(parts[4])

                    # The coordinates of the four points in the source image:
                    # top-left, top-right, bottom-right, and bottom-left.
                    src_points = np.float32([np.float32(coord_upper_left), np.float32(coord2_upper_right),
                                             np.float32(coord3_lower_right), np.float32(coord4_lower_left)])

                    # Coordinates of the four points in the destination image
                    dst_points = np.float32([[0, 0], [640, 0], [640, 640], [0, 640]])

                    # Calculate the perspective transformation matrix
                    M = cv2.getPerspectiveTransform(src_points, dst_points)

                    # Perform perspective transformation on the source image
                    dst_img = cv2.warpPerspective(im, M, (640, 640))

                    # Cropping images and add to img2predict for further detection
                    if mode == "test":
                        cv2.imwrite(f"evaluation/tmp/warp_img_{i}_{index}_{window_no}.png", dst_img)
                        self.img2predict.append(f"evaluation/tmp/warp_img_{i}_{index}_{window_no}.png")
                    if mode == "GUI":
                        cv2.imwrite(f"images/tmp/warp_img_{i}_{index}_{window_no}.png", dst_img)
                        self.img2predict.append(f"images/tmp/warp_img_{i}_{index}_{window_no}.png")
                    index += 1

            # Display images in GUI
            if mode == "GUI":
                im = cv2.cvtColor(im, cv2.COLOR_BGR2RGB)
                height, width, channels = im.shape
                bytesPerLine = channels * width
                qimage = QImage(im.data, width, height, bytesPerLine, QImage.Format_RGB888)
                pixmap = QPixmap.fromImage(qimage)
                self.ms.update_image.emit(pixmap, i)

    def detect(self, mode):
        imgsz = [640, 640]  # inference size (pixels)
        index = 0
        model = self.model
        output_size = self.output_size
        max_det = 10  # maximum detections per image
        view_img = False  # show results
        save_crop = False  # save cropped prediction boxes
        nosave = False  # do not save images
        classes = None  # filter by class: --class 0, or --class 0 2 3
        agnostic_nms = False  # class-agnostic NMS
        augment = False  # ugmented inference
        visualize = False  # visualize features
        line_thickness = 3  # bounding box thickness (pixels)
        hide_labels = False  # hide labels
        hide_conf = False  # hide confidences
        half = False  # use FP16 half-precision inference
        conf_thres = self.conf_thres
        iou_thres = self.iou_thres
        upper_thres = 100   # threshold for detecting upper blocked
        lower_thres = 500   # threshold for detecting lower blocked
        middle_thres = 50   # threshold for detecting middle blocked
        num_images = len(self.img2predict)

        # Iterate all cropped images
        for source in self.img2predict:
            # Initialize parameters
            glass_detected = False
            shutter_detected = False
            glass_upper = 640
            shutter_lower = 0
            window_no = self.window_no[index]
            button_no = f"{window_no}_Button"

            if source == "":
                QMessageBox.warning(None, "warning", "Please upload images before detection")
            else:
                source = str(source)
                device = select_device(self.device)
                stride, names, pt, jit, onnx = model.stride, model.names, model.pt, model.jit, model.onnx
                imgsz = check_img_size(imgsz, s=stride)  # check image size
                save_img = not nosave and not source.endswith('.txt')  # save inference images

                dataset = LoadImages(source, img_size=imgsz, stride=stride, auto=pt and not jit)

                # Run inference
                if pt and device.type != 'cpu':
                    model(torch.zeros(1, 3, *imgsz).to(device).type_as(next(model.model.parameters())))  # warmup
                dt, seen = [0.0, 0.0, 0.0], 0
                for path, im, im0s, vid_cap, s in dataset:
                    t1 = time_sync()
                    im = torch.from_numpy(im).to(device)
                    im = im.half() if half else im.float()  # uint8 to fp16/32
                    im /= 255  # 0 - 255 to 0.0 - 1.0
                    if len(im.shape) == 3:
                        im = im[None]  # expand for batch dim
                    t2 = time_sync()
                    dt[0] += t2 - t1
                    pred = model(im, augment=augment, visualize=visualize)
                    t3 = time_sync()
                    dt[1] += t3 - t2
                    # NMS
                    pred = non_max_suppression(pred, conf_thres, iou_thres, classes, agnostic_nms, max_det=max_det)
                    dt[2] += time_sync() - t3

                    for i, det in enumerate(pred):  # per image
                        seen += 1
                        p, im0, frame = path, im0s.copy(), getattr(dataset, 'frame', 0)
                        p = Path(p)  # to Path
                        s += '%gx%g ' % im.shape[2:]  # print string
                        annotator = Annotator(im0, line_width=line_thickness, example=str(names))
                        img_height = im0.shape[0]
                        if len(det):

                            # Rescale boxes from img_size to im0 size
                            det[:, :4] = scale_coords(im.shape[2:], det[:, :4], im0.shape).round()

                            # Print results
                            for c in det[:, -1].unique():
                                n = (det[:, -1] == c).sum()  # detections per class
                                s += f"{n} {names[int(c)]}{'s' * (n > 1)}, "  # add to string

                            # Write results
                            for *xyxy, conf, cls in reversed(det):

                                # Glass detected
                                if cls == 0:

                                    glass_detected = True
                                    glass_upper = min(glass_upper, int(xyxy[1]))

                                    if save_img or save_crop or view_img:  # Add bbox to image
                                        c = int(cls)  # integer class
                                        label = None if hide_labels else (
                                            names[c] if hide_conf else f'{names[c]} {conf:.2f}')
                                        annotator.box_label(xyxy, label, color=colors(c, True))

                                # Roller shutter detected
                                if cls == 1:

                                    shutter_detected = True
                                    shutter_lower = max(shutter_lower, int(xyxy[3]))

                                    if save_img or save_crop or view_img:  # Add bbox to image
                                        c = int(cls)  # integer class
                                        label = None if hide_labels else (
                                            names[c] if hide_conf else f'{names[c]} {conf:.2f}')
                                        annotator.box_label(xyxy, label, color=colors(c, True))

                            if glass_detected and not shutter_detected:

                                # preset threshold 100 for upper window frame
                                # if upper window frame is not smaller than 100 pixel, it is considered upper blocked
                                if glass_upper < upper_thres:
                                    # counter for open-close fusion
                                    self.windows[window_no].count_open += 1

                                    # fusion logic: case 5
                                    if self.windows[window_no].count_open > self.windows[window_no].count_closed:
                                        self.windows[window_no].status = "open"

                                else:
                                    # only when the previous status is not "open" or "closed",
                                    # we consider "blocked" cases in later detection
                                    if self.windows[window_no].status == "blocked" or self.windows[
                                        window_no].status == "not_detected":
                                        self.windows[window_no].status = "blocked"
                                        self.windows[window_no].doc_max = min(self.windows[window_no].doc_max,
                                                                              glass_upper * 100 / img_height)

                            if not glass_detected and shutter_detected:

                                glass_upper = img_height
                                # preset threshold 500 for lower window frame
                                # if lower window frame is not bigger than 500 pixel, we consider it lower blocked
                                if shutter_lower > lower_thres:
                                    # counter for open-close fusion
                                    self.windows[window_no].count_closed += 1

                                    # fusion logic: case 4
                                    if not self.windows[window_no].count_open > self.windows[
                                        window_no].count_closed:
                                        self.windows[window_no].status = "closed"
                                        if self.windows[window_no].doc == 0:
                                            self.windows[window_no].doc = max(self.windows[window_no].doc,
                                                                              shutter_lower * 100 / img_height)
                                        else:
                                            self.windows[window_no].doc = (self.windows[
                                                                               window_no].doc + shutter_lower * 100 / img_height) / 2

                                else:
                                    if self.windows[window_no].status == "blocked" or self.windows[
                                        window_no].status == "not_detected":
                                        self.windows[window_no].status = "blocked"
                                        self.windows[window_no].doc_min = max(self.windows[window_no].doc_min,
                                                                              shutter_lower * 100 / img_height)

                            if glass_detected and shutter_detected:
                                # preset threshold 50 for the position difference of shutters and window glass
                                # if it is not smaller than 50 pixel, we consider it middle blocked
                                if shutter_lower - glass_upper < middle_thres:

                                    self.windows[window_no].count_closed += 1

                                    # fusion logic: case 4
                                    if not self.windows[window_no].count_open > self.windows[window_no].count_closed:
                                        self.windows[window_no].status = "closed"
                                        if self.windows[window_no].doc == 0:
                                            self.windows[window_no].doc = max(self.windows[window_no].doc,
                                                                              shutter_lower * 100 / img_height)
                                        else:
                                            self.windows[window_no].doc = (self.windows[
                                                                               window_no].doc + shutter_lower * 100 / img_height) / 2

                                else:
                                    if self.windows[window_no].status == "blocked" or self.windows[
                                        window_no].status == "not_detected":
                                        self.windows[window_no].status = "blocked"
                                        self.windows[window_no].doc_min = max(self.windows[window_no].doc_min,
                                                                              shutter_lower * 100 / img_height)

                            # if not glass_detected and not shutter_detected:
                            else:
                                # Fusion logic: case 3
                                if self.windows[window_no].status == "blocked" or self.windows[
                                    window_no].status == "not_detected":
                                    self.windows[window_no].status = "blocked"
                                    self.windows[window_no].doc_min = max(self.windows[window_no].doc_min,
                                                                          shutter_lower * 100 / img_height)
                                    self.windows[window_no].doc_max = min(self.windows[window_no].doc_max,
                                                                          glass_upper * 100 / img_height)

                        else:
                            if self.windows[window_no].status == "blocked" or self.windows[
                                window_no].status == "not_detected":
                                self.windows[window_no].status = "blocked"

                        if mode == "GUI":
                            self.ms.update_status.emit(button_no, self.windows[window_no].status)

                        LOGGER.info(f'{s}Done. ({t3 - t2:.3f}s)')
                        # Stream results
                        im0 = annotator.result()
                        resize_scale = output_size / im0.shape[0]
                        im0 = cv2.resize(im0, (0, 0), fx=resize_scale, fy=resize_scale)
                        file_name, file_ext = os.path.splitext(os.path.basename(p))
                        result_file_name = window_no + "_result" + file_ext
                        i = 1
                        # Save result images
                        if mode == "test":
                            while os.path.exists(os.path.join("evaluation/result", result_file_name)):
                                result_file_name = window_no + f"_result_{i}" + file_ext
                                i += 1
                            cv2.imwrite(os.path.join("evaluation/result", result_file_name), im0)

                        if mode == "GUI":
                            while os.path.exists(os.path.join("images/result", result_file_name)):
                                result_file_name = window_no + f"_result_{i}" + file_ext
                                i += 1
                            cv2.imwrite(os.path.join("images/result", result_file_name), im0)
                            self.ms.update_processBar.emit((index + 1) * 100 // num_images)

                        index += 1

        if mode == "GUI":
            self.ms.update_graph.emit()

    def export(self):
        # create a new Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # add information
        sheet['A1'] = 'window_no'
        sheet['B1'] = 'status'
        sheet['C1'] = 'degree of closure(DoC%)'
        sheet['D1'] = 'max DoC (if blocked)'
        sheet['E1'] = 'min DoC (if blocked)'

        for i in range(1, 28):
            row_num = i + 1
            sheet['A{}'.format(row_num)] = 'upper_{}'.format(i)
            sheet['B{}'.format(row_num)] = self.windows['upper_{}'.format(i)].status
            sheet['C{}'.format(row_num)] = 0 if self.windows['upper_{}'.format(i)].status == "open" else \
                float(self.windows['upper_{}'.format(i)].doc) / 100 if \
                    self.windows['upper_{}'.format(i)].status == "closed" else ''
            sheet['D{}'.format(row_num)] = float(self.windows['upper_{}'.format(i)].doc_max) / 100 if \
                self.windows['upper_{}'.format(i)].status == "blocked" else ''
            sheet['E{}'.format(row_num)] = float(self.windows['upper_{}'.format(i)].doc_min) / 100 if \
                self.windows['upper_{}'.format(i)].status == "blocked" else ''

        for i in range(1, 28):
            row_num = i + 28 + 1
            sheet['A{}'.format(row_num)] = 'middle_{}'.format(i)
            sheet['B{}'.format(row_num)] = self.windows['middle_{}'.format(i)].status
            sheet['C{}'.format(row_num)] = 0 if self.windows['middle_{}'.format(i)].status == "open" else \
                float(self.windows['middle_{}'.format(i)].doc) / 100 if \
                    self.windows['middle_{}'.format(i)].status == "closed" else ''
            sheet['D{}'.format(row_num)] = float(self.windows['middle_{}'.format(i)].doc_max) / 100 if \
                self.windows['middle_{}'.format(i)].status == "blocked" else ''
            sheet['E{}'.format(row_num)] = float(self.windows['middle_{}'.format(i)].doc_min) / 100 if \
                self.windows['middle_{}'.format(i)].status == "blocked" else ''

        for i in range(1, 18):
            row_num = i + 56 + 1
            sheet['A{}'.format(row_num)] = 'lower_{}'.format(i)
            sheet['B{}'.format(row_num)] = self.windows['lower_{}'.format(i)].status
            sheet['C{}'.format(row_num)] = 0 if self.windows['lower_{}'.format(i)].status == "open" else \
                float(self.windows['lower_{}'.format(i)].doc) / 100 if \
                    self.windows['lower_{}'.format(i)].status == "closed" else ''
            sheet['D{}'.format(row_num)] = float(self.windows['lower_{}'.format(i)].doc_max) / 100 if \
                self.windows['lower_{}'.format(i)].status == "blocked" else ''
            sheet['E{}'.format(row_num)] = float(self.windows['lower_{}'.format(i)].doc_min) / 100 if \
                self.windows['lower_{}'.format(i)].status == "blocked" else ''

        for col in ['C', 'D', 'E']:
            for i in range(2, 84):
                cell = sheet['{}{}'.format(col, i)]
                if cell.value != '':
                    cell.number_format = '0.00%'

        # Modify table display style
        for col in sheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.font = Font(name='timesnewroman', size=11, bold=True)
        sheet.row_dimensions[1].height = 25

        for column_index in range(1, 6):
            header_cell = sheet.cell(row=1, column=column_index)
            header_cell.font = Font(name='Times New Roman', bold=True, size=11)
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.border = Border(bottom=Side(style='thin', color='000000'))

        for column in sheet.columns:
            column_length = max(len(str(cell.value)) for cell in column)
            column_letter = get_column_letter(column[0].column)
            sheet.column_dimensions[column_letter].width = column_length + 4

        color_map = {
            'open': '87CEEB',
            'closed': 'FFA500',
            'blocked': 'FF4500'
        }

        for i in range(1, 28):
            row_num = i + 1
            cell = sheet['B{}'.format(row_num)]
            status = self.windows['upper_{}'.format(i)].status
            if status in color_map:
                cell.fill = PatternFill(start_color=color_map[status], end_color=color_map[status], fill_type='solid')

        for i in range(1, 28):
            row_num = i + 29
            cell = sheet['B{}'.format(row_num)]
            status = self.windows['middle_{}'.format(i)].status
            if status in color_map:
                cell.fill = PatternFill(start_color=color_map[status], end_color=color_map[status], fill_type='solid')
        for i in range(1, 19):
            row_num = i + 57
            cell = sheet['B{}'.format(row_num)]
            status = self.windows['lower_{}'.format(i)].status
            if status in color_map:
                cell.fill = PatternFill(start_color=color_map[status], end_color=color_map[status], fill_type='solid')

        # Add charts to the table
        img1_pil = Image.open('images/result/bar_chart.png')
        img1 = XLImage(img1_pil)
        img1.width = 300
        img1.height = 220
        sheet.add_image(img1, 'G16')

        img2_pil = Image.open('images/result/pie_chart.png')
        img2 = XLImage(img2_pil)
        img2.width = 300
        img2.height = 220
        sheet.add_image(img2, 'G3')
        # Set default file name as the photo taken time

        self.workbook = workbook

    def save(self):

        window_info = self.window_info
        modified_status = self.modified_status
        window_no = window_info.split()[0]
        self.windows[window_no].status = modified_status

        if modified_status == "open":
            self.windows[window_no].doc = 0

        if modified_status == "closed":
            # Get doc from windowText displayed
            start = window_info.find("(") + 1
            end = window_info.find(")")
            try:
                new_doc = float(window_info[start:end].strip("%"))
            except Exception:
                QMessageBox.information(None, "Information",
                                        "DoC has been set to 100% (default), localizing is suggested before you save the changes!")
                new_doc = 100.0

            self.windows[window_no].doc = new_doc

        if modified_status == "blocked":
            # Get doc_min and doc_max from windowText displayed
            start = window_info.find("(") + 1
            end = window_info.find(")")
            try:
                new_doc_min = float(window_info[start:end].split()[0].strip("%"))
                new_doc_max = float(window_info[start:end].split()[2].strip("%"))
            except Exception:
                QMessageBox.information(None, "Information",
                                        "DoC has been set to (0-100%) (default), localizing is suggested before you save the changes!")
                new_doc_min = 0.0
                new_doc_max = 100.0

            self.windows[window_no].doc_min = new_doc_min
            self.windows[window_no].doc_max = new_doc_max

        self.ms.update_windowInfo.emit(window_no, modified_status)
        self.ms.update_graph.emit()
